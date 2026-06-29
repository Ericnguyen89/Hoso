# -*- coding: utf-8 -*-
"""
Tạo bìa hồ sơ lưu trữ (.docx) hàng loạt từ dữ liệu Excel (.xlsx).

Cách hoạt động
--------------
- Dùng chính tệp BÌA mẫu (.docx) làm khuôn: giữ NGUYÊN khổ A4, lề, text box,
  viền khung, font, cỡ chữ và các tab stop của mẫu.
- Mỗi dòng dữ liệu trong Excel -> 1 bìa, mỗi bìa nằm trên 1 trang A4 riêng.
- Chỉ các GIÁ TRỊ thay đổi (Đảng bộ, tiêu đề, ngày, số trang, ...) được thay,
  toàn bộ định dạng được sao chép từ mẫu.

Cách dùng
---------
    python tao_bia.py
hoặc tùy biến đường dẫn:
    python tao_bia.py --mau "BÌA VV 14-19.docx" --data "Dữ liệu bìa.xlsx" --out "Bia_ho_so.docx"

Cấu trúc cột Excel (Sheet đầu tiên, không có dòng tiêu đề):
    A: Hồ sơ số            (vd 0001)
    B: Tiêu đề hồ sơ
    C: Thời gian           (vd 16/02/2006-19/02/2008)  -> Bắt đầu / Kết thúc
    D: Số trang
    E: Số TL (số tờ)
    F: Thời hạn bảo quản   (vd Vĩnh viễn)
    G: Dòng Đảng bộ (đầu bìa)
"""

import argparse
import copy
import os
import re
import sys

from lxml import etree
import openpyxl

W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"
A = "http://schemas.openxmlformats.org/drawingml/2006/main"
WPS = "http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
WP = "http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"
WP14 = "http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing"


def w(tag):
    return "{%s}%s" % (W, tag)


def qn(ns, tag):
    return "{%s}%s" % (ns, tag)


# ----- Giá trị MẪU (lấy từ bìa mẫu, ứng với hồ sơ 0001) dùng để định vị run -----
# Mỗi giá trị dưới đây là một run riêng biệt, duy nhất trong text box mẫu.
MAU = {
    "org": "PHÔNG LƯU TRỮ CƠ QUAN LÃNH ĐẠO XÃ TAM ĐẠI, KHÓA XIX, "
           "NHIỆM KỲ 2005-2010",
    "title": "Tập Nghị quyết của Đảng ủy xã Tam Đại về thực hiện các nhiệm vụ "
             "chính trị. Năm: 2006-2008",
    "start": "16/02/2006",
    "end": "19/02/2008",
    "ho_so_so": "0001",
    "so_trang": "61",
    "so_tl": "16",
    "thbq": "Vĩnh viễn",
}


def runs_text(p):
    """Văn bản hiển thị của 1 đoạn (gộp các <w:t>)."""
    return "".join(t.text or "" for t in p.iter(w("t")))


def set_run_text(run, text):
    """Đặt lại văn bản cho 1 run, giữ nguyên định dạng (rPr) và xml:space."""
    # Xóa các node text/tab cũ, thêm 1 <w:t> mới (giữ rPr ở đầu)
    for child in list(run):
        if child.tag in (w("t"), w("tab")):
            run.remove(child)
    t = etree.SubElement(run, w("t"))
    t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
    t.text = text


def replace_in_paragraph(p, old, new):
    """Thay run đầu tiên trong đoạn p có văn bản đúng bằng `old`."""
    for r in p.findall(w("r")):
        if "".join(t.text or "" for t in r.iter(w("t"))) == old:
            set_run_text(r, new)
            return True
    return False


def fill_textbox(txbx, row):
    """Điền dữ liệu cho 1 text box (1 bìa).

    Định vị theo NỘI DUNG (giá trị mẫu) chứ không theo số thứ tự đoạn, nên
    bền vững với việc mẫu thêm/bớt dòng. Các giá trị mẫu trong MAU đều là
    những run riêng biệt và duy nhất trong text box.
    """
    # (giá trị_mẫu -> giá trị_mới); thứ tự không quan trọng vì ta CLAIM run
    # theo văn bản gốc của mẫu trước, rồi mới ghi đè.
    mapping = [
        (MAU["org"], row["org"]),
        (MAU["title"], row["title"]),
        (MAU["start"], row["start"]),
        (MAU["end"], row["end"]),
        (MAU["ho_so_so"], row["ho_so_so"]),
        (MAU["so_trang"], row["so_trang"]),
        (MAU["so_tl"], row["so_tl"]),
        (MAU["thbq"], row["thbq"]),
    ]

    # Liệt kê tất cả run trong text box kèm văn bản hiện tại
    runs = [(r, "".join(t.text or "" for t in r.iter(w("t"))))
            for r in txbx.iter(w("r"))]

    claimed = set()          # id(run) đã được gán
    todo = []                # (run, giá_trị_mới)
    for old, new in mapping:
        for r, txt in runs:
            if id(r) in claimed:
                continue
            if txt == old:
                claimed.add(id(r))
                todo.append((r, new))
                break
    for r, new in todo:
        set_run_text(r, new)


def normalize_date(token):
    """Chuẩn hóa 1 mốc ngày:
       - Đủ ngày/tháng/năm -> dd/mm/yyyy (ngày & tháng 2 chữ số, năm 4 chữ số)
       - Chỉ tháng/năm (vd '8/2019') -> mm/yyyy (tháng 2 chữ số, năm 4 chữ số)
       - Chỉ năm           -> yyyy
       - Dạng dính liền 'ddmmyyyy' (vd '22102020') -> dd/mm/yyyy
    """
    s = (token or "").strip().replace(" ", "")
    if not s:
        return ""

    def y4(y):
        return y if len(y) >= 4 else y.zfill(4)

    # 8 chữ số dính liền: ddmmyyyy
    if re.fullmatch(r"\d{8}", s):
        return "%s/%s/%s" % (s[0:2], s[2:4], s[4:8])

    parts = [p for p in s.split("/") if p != ""]
    if len(parts) == 3:                       # ngày / tháng / năm
        d, m, y = parts
        return "%s/%s/%s" % (d.zfill(2), m.zfill(2), y4(y))
    if len(parts) == 2:                       # tháng / năm (không có ngày)
        m, y = parts
        return "%s/%s" % (m.zfill(2), y4(y))
    if len(parts) == 1:                       # chỉ năm
        return parts[0]
    return s


def split_dates(raw):
    """'16/02/2006-19/02/2008' -> ('16/02/2006', '19/02/2008').
    Chuẩn hóa định dạng; chịu được khoảng trắng và trường hợp chỉ có 1 ngày."""
    s = (raw or "").strip()
    if not s:
        return "", ""
    # Tách trên dấu '-' nằm GIỮA hai ngày (ngày dạng dd/mm/yyyy không chứa '-')
    parts = re.split(r"\s*-\s*", s)
    if len(parts) >= 2:
        return normalize_date(parts[0]), normalize_date(parts[-1])
    return normalize_date(s), ""


def cell(ws, r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()


def is_header(a, c, d):
    """Dòng tiêu đề: không có chữ số ở các cột Hồ sơ số / Thời gian / Số trang."""
    return not re.search(r"\d", a + " " + c + " " + d)


def read_rows(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    return read_rows_ws(wb.worksheets[0])


def read_rows_ws(ws):
    rows = []
    for r in range(1, ws.max_row + 1):
        a = cell(ws, r, 1)   # hồ sơ số
        b = cell(ws, r, 2)   # nội dung (tiêu đề hồ sơ)
        c = cell(ws, r, 3)   # thời gian (bắt đầu-kết thúc)
        d = cell(ws, r, 4)   # số trang
        e = cell(ws, r, 5)   # số TL (số tệp)
        f = cell(ws, r, 6)   # thời hạn bảo quản
        g = cell(ws, r, 7)   # dòng phông/đảng bộ (đầu bìa)
        # Bỏ qua dòng trống hoàn toàn và dòng tiêu đề
        if not any([a, b, c, d, e, f, g]):
            continue
        if is_header(a, c, d):
            continue
        start, end = split_dates(c)
        rows.append({
            "ho_so_so": a,
            "title": b,
            "start": start,
            "end": end,
            "so_trang": d,
            "so_tl": e,
            "thbq": f,
            "org": g,
        })
    return rows


def _bottom_block_index(txbx):
    """Chỉ số đoạn bắt đầu khối dưới (đoạn chứa 'Phông')."""
    for i, p in enumerate(txbx.findall(w("p"))):
        if "Phông" in "".join(t.text or "" for t in p.iter(w("t"))):
            return i
    return None


def _set_nofill(sppr):
    """Bỏ nền và viền của shape (đặt noFill)."""
    for tag in ("solidFill", "gradFill", "blipFill", "pattFill", "noFill"):
        e = sppr.find(qn(A, tag))
        if e is not None:
            sppr.remove(e)
    # Chèn noFill ngay sau prstGeom (đúng thứ tự schema)
    geom = sppr.find(qn(A, "prstGeom"))
    nofill = etree.Element(qn(A, "noFill"))
    if geom is not None:
        geom.addnext(nofill)
    else:
        sppr.append(nofill)
    ln = sppr.find(qn(A, "ln"))
    if ln is not None:
        for c in list(ln):
            ln.remove(c)
        etree.SubElement(ln, qn(A, "noFill"))


def restructure_cover(cover_p):
    """Tách bìa thành 2 text box phủ toàn trang chồng lên nhau:
       - Box 1: giữ khung viền + nền + nội dung trên (đến trước 'Phông số'),
         canh đỉnh (anchor='t').
       - Box 2: trong suốt, không viền, chỉ chứa khối dưới ('Phông số'...),
         canh ĐÁY (anchor='b') -> luôn cố định ở đáy trang.
    """
    run = cover_p.find(w("r"))
    ac = run.find(qn(MC, "AlternateContent"))
    if ac is None:
        return  # đã tái cấu trúc rồi hoặc mẫu khác

    # 1) Bỏ AlternateContent, chỉ giữ <w:drawing> trong mc:Choice (box 1)
    choice = ac.find(qn(MC, "Choice"))
    draw1 = copy.deepcopy(choice.find(w("drawing")))
    idx = list(run).index(ac)
    run.remove(ac)
    run.insert(idx, draw1)

    txbx1 = draw1.find(".//" + w("txbxContent"))
    bidx = _bottom_block_index(txbx1)
    if bidx is None:
        return

    # 2) Tạo box 2 từ bản sao box 1
    draw2 = copy.deepcopy(draw1)

    # Box 1: bỏ các đoạn từ 'Phông số' trở xuống
    for p in txbx1.findall(w("p"))[bidx:]:
        txbx1.remove(p)

    # Box 2: chỉ giữ các đoạn từ 'Phông số' trở xuống
    txbx2 = draw2.find(".//" + w("txbxContent"))
    for p in txbx2.findall(w("p"))[:bidx]:
        txbx2.remove(p)

    # Box 2: canh nội dung xuống đáy
    body2 = draw2.find(".//" + qn(WPS, "bodyPr"))
    if body2 is not None:
        body2.set("anchor", "b")
    # Box 2: bỏ nền + viền (để lộ khung của box 1)
    sp2 = draw2.find(".//" + qn(WPS, "spPr"))
    if sp2 is not None:
        _set_nofill(sp2)
    # Box 2: id/anchorId/z-order khác box 1 để hợp lệ và nằm trên
    anc2 = draw2.find(qn(WP, "anchor"))
    if anc2 is not None:
        anc2.set(qn(WP14, "anchorId"), "0CBB7FA1")
        anc2.set(qn(WP14, "editId"), "040BC5A1")
        rh = anc2.get("relativeHeight")
        if rh:
            anc2.set("relativeHeight", str(int(rh) + 4))
    doc2 = draw2.find(".//" + qn(WP, "docPr"))
    if doc2 is not None:
        try:
            doc2.set("id", str(int(doc2.get("id")) + 1))
        except (TypeError, ValueError):
            doc2.set("id", "777000001")
        doc2.set("name", "BiaDuoi")

    # 3) Thêm box 2 vào cùng đoạn (run mới) -> hai box phủ cùng vị trí
    run2 = etree.SubElement(cover_p, w("r"))
    run2.append(draw2)


def build_document_xml(doc_xml, rows, progress=None):
    """Từ document.xml của mẫu + danh sách dòng -> document.xml mới (bytes).
    progress(done, total): hàm tùy chọn báo tiến độ sau mỗi bìa."""
    root = etree.fromstring(doc_xml)
    body = root.find(w("body"))

    # Tách các thành phần body: đoạn chứa text box (cover) và sectPr
    children = list(body)
    cover_p = None
    sectpr = None
    for ch in children:
        if ch.tag == w("p") and ch.find(".//" + w("drawing")) is not None:
            cover_p = ch
        elif ch.tag == w("sectPr"):
            sectpr = ch
    if cover_p is None or sectpr is None:
        raise ValueError("Không tìm thấy text box bìa hoặc sectPr trong tệp mẫu.")

    # Cố định khối dưới ở đáy trang (tách thành text box thứ 2 canh đáy)
    restructure_cover(cover_p)

    # Dọn sạch body, dựng lại: cover_1, [ngắt trang], cover_2, ..., sectPr
    for ch in children:
        body.remove(ch)

    for idx, row in enumerate(rows):
        cp = copy.deepcopy(cover_p)
        # Điền dữ liệu cho TẤT CẢ text box bên trong (Choice + Fallback)
        for txbx in cp.iter(w("txbxContent")):
            fill_textbox(txbx, row)
        # Từ bìa thứ 2: ngắt sang trang mới
        if idx > 0:
            ppr = cp.find(w("pPr"))
            if ppr is None:
                ppr = etree.Element(w("pPr"))
                cp.insert(0, ppr)
            etree.SubElement(ppr, w("pageBreakBefore"))
        body.append(cp)
        if progress is not None:
            progress(idx + 1, len(rows))

    body.append(sectpr)
    return etree.tostring(root, xml_declaration=True,
                          encoding="UTF-8", standalone=True)


def assemble_docx(mau_bytes, new_document_xml):
    """Sao chép toàn bộ docx mẫu (bytes), chỉ thay word/document.xml -> bytes."""
    import zipfile, io
    src = io.BytesIO(mau_bytes)
    out = io.BytesIO()
    with zipfile.ZipFile(src) as zin, \
            zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "word/document.xml":
                data = new_document_xml
            zout.writestr(item, data)
    return out.getvalue()


def generate_from_bytes(mau_bytes, data_bytes, progress=None):
    """Lõi dùng cho web: nhận bytes mẫu + bytes Excel -> (docx bytes, danh sách dòng).
    progress(done, total): hàm tùy chọn báo tiến độ realtime."""
    import io, zipfile
    rows = read_rows_ws(
        openpyxl.load_workbook(io.BytesIO(data_bytes), data_only=True).worksheets[0])
    if not rows:
        raise ValueError("Không đọc được dòng dữ liệu nào từ Excel.")
    doc_xml = zipfile.ZipFile(io.BytesIO(mau_bytes)).read("word/document.xml")
    new_doc = build_document_xml(doc_xml, rows, progress=progress)
    return assemble_docx(mau_bytes, new_doc), rows


def build(mau_path, data_path, out_path):
    with open(mau_path, "rb") as f:
        mau_bytes = f.read()
    with open(data_path, "rb") as f:
        data_bytes = f.read()
    out_bytes, rows = generate_from_bytes(mau_bytes, data_bytes)
    with open(out_path, "wb") as f:
        f.write(out_bytes)
    print("Đã tạo %d bìa -> %s" % (len(rows), out_path))


def autofind(here, pattern, exclude=()):
    import glob
    cands = [p for p in glob.glob(os.path.join(here, pattern))
             if os.path.basename(p) not in exclude]
    return cands[0] if cands else None


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="Tạo bìa hồ sơ từ Excel.")
    ap.add_argument("--mau", default=None, help="Tệp .docx mẫu (mặc định: tự dò)")
    ap.add_argument("--data", default=None, help="Tệp .xlsx dữ liệu (mặc định: tự dò)")
    ap.add_argument("--out", default=os.path.join(here, "Bia_ho_so_ket_qua.docx"),
                    help="Tệp .docx kết quả")
    args = ap.parse_args()

    # Tự dò tệp nếu không chỉ định (tránh nhầm tệp kết quả)
    mau = args.mau or autofind(here, "*.docx",
                               exclude={os.path.basename(args.out)})
    data = args.data or autofind(here, "*.xlsx")

    if not mau or not os.path.exists(mau):
        sys.exit("Không tìm thấy tệp .docx mẫu.")
    if not data or not os.path.exists(data):
        sys.exit("Không tìm thấy tệp .xlsx dữ liệu.")

    print("Mẫu :", os.path.basename(mau))
    print("Data:", os.path.basename(data))
    build(mau, data, args.out)


if __name__ == "__main__":
    main()
